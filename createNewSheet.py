import sys
import openpyxl
import datetime
import re
import os
import datetime
import logging
import servicenowRestAPI
import random
import urllib
import createPartnerSheet



from datetime import datetime,timedelta,date
from openpyxl import load_workbook
from datetime import datetime
from operator import itemgetter, attrgetter, methodcaller
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle, Font, Border,Side, PatternFill, Font, GradientFill, Alignment, colors
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import BORDER_NONE


varTitle = '| RedBrick Health'
conferenceInfo = ''  # Not used
meetingInfo = ''  # Not used
yellowrange = 10
redrange = -1
logo_url ='http://image.get.healthyemail.com/lib/fe8e13727c62047c7d/m/1/4ee4b6bc-ace4-4829-9d7b-079c044fcf28.png?b=1516821132000'

lightColor = '20 % - Accent1'
darkerColor = '40 % - Accent1'


wsClientAction = 'Client Action Items'
wsAccountPlanning = 'Account Planning'
wsClosedOperations = 'Closed Operations'
wsClientActionTabColor = 'ff7f7f'
# global across each meeting record
Waiting_for_Account_Tasks = []


def updateExcelwithcases(url_servicenow,user,pwd,wb,newFilepath,caselist,accountname,account_sys_id):


    #TODO take list and make sublists to post into excel

    closed_cases = []
    Account_Planning = []
    Consumer_Inquiry = []
    Account_Strategy = []
    Account_Request = []
    Customer_Inquiry = []
    Account_Objective = []
    Waiting_for_Account = []
    global Waiting_for_Account_Tasks
    Waiting_for_Account_Tasks = []

    left_overs = []  # Should be empty

    logging.info('Processing cases into categories')

    for s in caselist:


        if  s.get("active") == "false":
            closed_cases.append(s)
        elif s.get("category") == "Account Planning":
            Account_Planning.append(s)
        elif s.get("category") == "Account Requests":
            Account_Request.append(s)
        elif s.get("category") == "Consumer Inquiry":
            Consumer_Inquiry.append(s)
        elif s.get("category") == "Account Strategy":
            Account_Strategy.append(s)
        elif s.get("category") == "Customer Inquiry":
            Customer_Inquiry.append(s)
        elif s.get("category") == "Account Objective":
            Account_Objective.append(s)

        if s.get("state") == "Awaiting Info":
            Waiting_for_Account.append(s)



    #TODO create header template row for each Category
    endrow  = '' # used to keep tracker where on the sheet

    #Account Plannin
    endrow = 4 # for starting after the title
    if len(Account_Planning) != 0:
        endrow  =  section_table(accountname, Account_Planning,wsAccountPlanning,endrow,wb,wsAccountPlanning,url_servicenow,user,pwd)
        logging.info('Account_Planning: Processing cases section_table')

    #Start Operations
    endrow = 4 # for starting after the title
    if len(Consumer_Inquiry) != 0:
        endrow  =  section_table(accountname,Consumer_Inquiry,'Consumer Inquiry',endrow,wb,'Operations',url_servicenow,user,pwd)
        logging.info('Consumer_Inquiry: Processing cases section_table')

    if len(Customer_Inquiry) != 0:
        endrow  =  section_table(accountname,Customer_Inquiry,'Customer Inquiry',endrow,wb,'Operations',url_servicenow,user,pwd)
        logging.info('Customer_Inquiry: Processing cases section_table')

    #Start Closed Operatations sheet
    endrow = 4 # for starting after the title
    if len(closed_cases) != 0:
        endrow  =  section_table(accountname,closed_cases,wsClosedOperations,endrow,wb,wsClosedOperations,url_servicenow,user,pwd)
        logging.info('closed_cases: Processing cases section_table')


    #Start Account Summary Sheet
    endrow = 4 # for starting after the title
    if len(Waiting_for_Account) != 0:
        endrow  =  section_table(accountname,Waiting_for_Account,wsClientAction,endrow,wb,wsClientAction,url_servicenow,user,pwd)
        logging.info('Waiting_for_Account: Processing cases section_table' + wsClientAction)

    #global Waiting_for_Account_Tasks
    if len(Waiting_for_Account_Tasks) != 0:
        #Create a gap between cases and tasks
        print Waiting_for_Account_Tasks
        endrow = endrow + 2
        endrow  =  section_tasktable(accountname,Waiting_for_Account_Tasks,'Client Open Tasks',endrow,wb,wsClientAction,url_servicenow,user,pwd)
        logging.info('Waiting_for_Account: Processing WAITING FOR TASKS section_table' + wsClientAction)

    if len(left_overs) != 0:
        endrow  =  section_table(accountname,left_overs,'Other Cases',endrow,wb,wsClosedOperations,url_servicenow,user,pwd)

    return wb

def checkworkbook(wb,accountname):

    #TODO WB Add Styles

    if wsAccountPlanning not in wb.sheetnames:
        createPlanningWSHeader(wb,wsAccountPlanning,accountname, conferenceInfo ,meetingInfo,2,7)
        logging.info("Created Sheet:" + wsAccountPlanning)
    if 'Operations' not in wb.sheetnames:
        createPlanningWSHeader(wb,'Operations',accountname, conferenceInfo ,meetingInfo,3,7)
        logging.info("Created Operations worksheet")

    if 'Closed Operations' not in wb.sheetnames:
        createPlanningWSHeader(wb,'Closed Operations',accountname, '' ,'',4,7)
        logging.info("Create Closed Items")

    if wsClientAction not in wb.sheetnames:
        createPlanningWSHeader(wb,wsClientAction,accountname, '' ,'',1,7)
        logging.info('Creating Sheet: ' + wsClientAction)

    return True

def createPlanningWSHeader (wb,name,accountname,meetingInfo,conferenceInfo,wb_order,totalColumns):
    logging.info('IN CreatePlanning to create:' + name +'Location' + str(wb_order))

    ws = wb.create_sheet(name, wb_order)
    ws.merge_cells('A1:I1')
    ws_title_cell = ws.cell(row=1, column=1)
    newtitle = str(accountname) + str(varTitle)
    ws_title_cell.value = newtitle
    ws_title_cell.border = BORDER_NONE
    ws_title_cell.style = 'Title'
    ws_title_cell.alignment = Alignment(horizontal='center',vertical='top',shrink_to_fit=True)

    ws.merge_cells('A2:I2')
    ws_conferneceInfo =  ws.cell(row=2, column=1)
    ws_conferneceInfo.value = conferenceInfo
    ws_conferneceInfo.style = 'Headline 3'
    ws_title_cell.border = BORDER_NONE
    ws_conferneceInfo.alignment = Alignment(wrap_text= True,horizontal='center',vertical='top',shrink_to_fit=True)

    ws.merge_cells('A3:I3')
    ws_meetingInfo =  ws.cell(row=3, column=1)
    ws_meetingInfo.value = meetingInfo
    ws_title_cell.border = BORDER_NONE
    ws_meetingInfo.style = 'Headline 3'
    ws_meetingInfo.alignment = Alignment(wrap_text= True,horizontal='center',vertical='top',shrink_to_fit=True)

    worksheetstyleguide(ws)


    logo_file = get_logoforHeader(logo_url)
    img = Image(logo_file)
    ws.add_image(img,'A1')


    return True

# make the worksheet standard
def worksheetstyleguide(ws):
    ws.fitToPage= True
    ws.tabColor = colors.BLUE
    ws.ORIENTATION_LANDSCAPE = 'landscape'
    ws.PAPERSIZE_LETTER = '1'
    # Might need to be adjusted to the per page
    ws.print_title_rows = '1:3'  # NOTE this was depricated, old code was ws1.add_print_title(2)


def createSheetProcesser(url_servicenow,user,pwd,newFilepath,accountname,account_sys_id):

    logging.info('Load workbook:' + newFilepath)
    wb = load_workbook(newFilepath)

    # Validate the excel document has the right tabs to process further.
    logging.info('Check workbook:' + newFilepath)
    checkworkbook(wb,accountname)

    logging.info('Get case list to populate sheet: ' + newFilepath)

    caselist = servicenowRestAPI.getcasesTOupdateExcel(url_servicenow,user,pwd,account_sys_id)

    logging.info('After case list to populate sheet: ' + newFilepath)

    updateExcelwithcases(url_servicenow,user,pwd,wb,newFilepath,caselist,accountname,account_sys_id)

    logging.info('After updateExcelwithcases: this step takes in all the cases and processes them into sheets')

    logging.info('Before checkifPartner:' + accountname )

    createPartnerSheet.checkifPartner(wb,accountname,newFilepath,url_servicenow,account_sys_id,user,pwd)

    logging.info('before saving file:' + newFilepath )

    wb = wb.save(newFilepath)

    logging.info('We saving the file:' + newFilepath )

    return newFilepath

def getdatetimeColorFormat(vdatetime,rowcolor):
    if type(vdatetime) is date:
        if vdatetime <= (datetime.now().date() - timedelta(days=redrange)):  # Red
            #return '20 % - Accent2'
            return 'Bad'
        elif vdatetime <= (datetime.now().date() - timedelta(days=yellowrange)):  # Yellow
            #return '20 % - Accent6'
            return 'Neutral'
    return rowcolor

def getWAITINGFORColorFormat(state ,rowcolor):
    if state == 'Awaiting Info':
        return 'Bad'
    if state == 'Parking Lot':
        return 'Neutral'
    return rowcolor
def getWAITINGFORaddCompanyName(accountname,state):
    if state == 'Awaiting Info':
        return 'Awaiting Info From ' + accountname
    return state



def get_logoforHeader (url):
    image_name = random.randrange(100,9999)
    full_name = 'redbrick_' +str(image_name) + '.png'
    try:
        urllib.urlretrieve(url,full_name)
    except TypeError:
        logging.error('Logo url is probably wrong or disk errors: '+url,1)
        raise
    finally:    
            return full_name

def section_tasktable(accountname,tasklist,title,rowStart,wb,sheetname,url_servicenow,user,pwd):
        logging.info('In section_tasktable: ' + accountname + ' | ' +  title + ' | ' + sheetname)
        logging.info(len(tasklist))
        highlights = lightColor

      # should create a class here for creating sheets but next time?
        ws = wb[sheetname]
        rowStartHolder = rowStart  # used to create a table
        celltitle = ws['A' + str(rowStart)]
        celltitle.value = title
        celltitle.style = 'Headline 1'

        celltitle = ws['A' + str(rowStart)]
        celltitle.value = title + ' | ' + accountname

        #ws.merge_cells('A2:D2')
        ws.merge_cells('A' + str(rowStart)+ ':i' +str(rowStart))

        rowStart = rowStart+1

        # add column headings. NB. these must be strings

        ws['a' + str(rowStart)] =  ("Number")
        ws['b' + str(rowStart)] =  "Short Description"
        ws['c' + str(rowStart)] =  "Priorty"
        ws['d' + str(rowStart)] =  "State"
        ws['e' + str(rowStart)] =  "Created"
        ws['f' + str(rowStart)] =  "Due Date"
        ws['g' + str(rowStart)] =  "RedBrick Owner"
        ws['h' + str(rowStart)] =  "Description"
        #ws['i' + str(rowStart)] =  "Next Step"


        for s in tasklist:
             
            rowStart = rowStart + 1

            if highlights == lightColor:
                highlights = darkerColor
            elif highlights == darkerColor:
                highlights = lightColor
            else:
                highlights = lightColor


            _cellA = ws['A' + str(rowStart)]
            _cellA.style = highlights
            v_hyperlink = url_servicenow + '/nav_to.do?uri=sn_customerservice_case.do?sys_id=' + s.get("u_case").get("value") + '%26sysparm_view=case'
                   
            _cellA.hyperlink = v_hyperlink
            _cellA.value = s.get("u_case").get("display_value") + '| \n' + s.get("number").get("display_value")
            _cellA.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

            _cellC = ws['B' + str(rowStart)]
            _cellC.style =  highlights
           # _cellC.value = item.get("short_description").get("display_value") + ' | \n' + s.get("short_description")
            _cellC.value = s.get("short_description").get("display_value")
            _cellC.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

            _cellD = ws['C' + str(rowStart)]
            _cellD.style =  highlights
            pValue = s.get("priority").get("display_value")
            pValue = pValue[3:]
            _cellD.value = pValue
            
            _cellD.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

            _cellE = ws['D' + str(rowStart)]
            _cellE.style =  highlights
            _cellE.style =  getWAITINGFORColorFormat(s.get("state").get("display_value"),highlights)
            _cellE.value = getWAITINGFORaddCompanyName(accountname,s.get("state").get("display_value"))
            _cellE.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)                

            sys_created_on_value = datetime.strptime(s.get("sys_created_on").get("display_value"), "%Y-%m-%d %H:%M:%S")
            _cellD = ws['E' + str(rowStart)]
            _cellD.style =  highlights
            _cellD.value = sys_created_on_value
            _cellD.number_format = 'MM/DD/YYYY'
            _cellD.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

            if isinstance(s.get("due_date").get("display_value"), datetime) == 'True':
                due_date_value = datetime.strptime(s.get("due_date").get("display_value"), "%Y-%m-%d %H:%M:%S")
            else:
                due_date_value = ''

            due_date_value = due_date_value
            _cellF = ws['F' + str(rowStart)]
            _cellF.style =  getdatetimeColorFormat(due_date_value,highlights)
            _cellF.value =  due_date_value
            _cellF.number_format = 'MM/DD/YYYY'
            _cellF.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

            _cellH = ws['G' + str(rowStart)]
            _cellH.style =  highlights
            _cellH.value = s.get("assigned_to").get("display_value")
            _cellH.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

            _cellI = ws['H' + str(rowStart)]
            _cellI.style = highlights
            _cellI.value = s.get("description").get("display_value")
            _cellI.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)



        # finished loop now create table and tab formatt            
        rowend = rowStart      
        refTables = 'A' + str(rowStartHolder+1)+ ':H' +str(rowend)
        
        tableName = title.replace(' ','')
        tab = Table(displayName= tableName, ref= refTables)

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    
        return rowend
    

def section_table (accountname,caselist,title,rowStart,wb,sheetname,url_servicenow,user,pwd):
    logging.info('In Section table: ' + accountname + ' | ' +  title + ' | ' + sheetname)
    highlights = lightColor

    # should create a class here for creating sheets but next time?
    ws = wb[sheetname]
    rowStartHolder = rowStart  # used to create a table
    celltitle = ws['A' + str(rowStart)]
    celltitle.value = title
    celltitle.style = 'Headline 1'

    celltitle = ws['A' + str(rowStart)]
    celltitle.value = title + ' | ' + accountname

    #ws.merge_cells('A2:D2')
    ws.merge_cells('A' + str(rowStart)+ ':i' +str(rowStart))

    rowStart = rowStart+1

    # add column headings. NB. these must be strings

    ws['a' + str(rowStart)] =  ("Number")
    ws['b' + str(rowStart)] =  "Short Description"
    ws['c' + str(rowStart)] =  "Priorty"
    ws['d' + str(rowStart)] =  "State"
    ws['e' + str(rowStart)] =  "Created"
    ws['f' + str(rowStart)] =  "Due Date"
    ws['g' + str(rowStart)] =  "RedBrick Owner"
    ws['h' + str(rowStart)] =  "Description/Completed Steps"
    ws['i' + str(rowStart)] =  "Next Step"

    if sheetname ==  wsClosedOperations:
        ws['j' + str(rowStart)] =  "Resolution"

    logging.info('Count AFTER: ROWS TITLES' + str(rowStart))
    logging.info('Before Addin: ROWS TITLES' + str(rowStart))

    for item in caselist:
        # added one for the title above and all the cases after
        rowStart = rowStart + 1
        
        if highlights == lightColor:
            highlights = darkerColor
        elif highlights == darkerColor:
            highlights = lightColor
        else:
            highlights = lightColor

        trackcolorissue = 'trackhighlight color issues: CurrentColor = ' + highlights + ' case= ' +  item.get("number") + ' rowID= ' + str(rowStart)
        logging.info(trackcolorissue)

        _cellA = ws['A' + str(rowStart)]
        _cellA.style = highlights
        v_hyperlink = url_servicenow + '/nav_to.do?uri=sn_customerservice_case.do?sys_id=' + item.get("sys_id") + '%26sysparm_view=case'
        _cellA.hyperlink = v_hyperlink
        _cellA.value = item.get("number")
        _cellA.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)


        _cellC = ws['B' + str(rowStart)]
        _cellC.style =  highlights
        _cellC.value = item.get("short_description")
        _cellC.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

        _cellJ = ws['C' + str(rowStart)]
        _cellJ.style =  highlights
        pValue = item.get("priority")
        pValue = pValue[3:]
        _cellJ.value =  pValue
        _cellJ.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

        _cellE = ws['D' + str(rowStart)]
        _cellE.style =  getWAITINGFORColorFormat(item.get("state"),highlights)
        _cellE.value = getWAITINGFORaddCompanyName(accountname,item.get("state"))
        _cellE.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

        sys_created_on_value = datetime.strptime(item.get("sys_created_on"), "%Y-%m-%d %H:%M:%S")
        _cellD = ws['E' + str(rowStart)]
        _cellD.style =  highlights
        _cellD.value = sys_created_on_value
        _cellD.number_format = 'MM/DD/YYYY'
        _cellD.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)


        due_date_value = datetime.strptime(item.get("due_date"), "%Y-%m-%d %H:%M:%S")
        _cellF = ws['F' + str(rowStart)]
        _cellF.style =  getdatetimeColorFormat(datetime.date(due_date_value),highlights)
        _cellF.value =  due_date_value
        _cellF.number_format = 'MM/DD/YYYY'
        _cellF.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

        _cellH = ws['G' + str(rowStart)]
        _cellH.style =  highlights
        #_cellH.value = item.get("assigned_to").get("display_value")
        _cellH.value = item.get("assigned_to").get("display_value") if hasattr(item.get("assigned_to"), 'get') else ""
        _cellH.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

        descValue = item.get("description") + '\n ' +  item.get("u_completed_steps")
        _cellH = ws['H' + str(rowStart)]
        _cellH.style = highlights
        _cellH.value = descValue
        _cellH.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

        _cellL = ws['I' + str(rowStart)]
        _cellL.style =  highlights
        _cellL.value = item.get("u_next_steps")
        _cellL.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

        if sheetname ==  wsClosedOperations:
            _cellL = ws['J' + str(rowStart)]
            _cellL.style =  highlights
            _cellL.value = item.get("close_notes")
            _cellL.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

        casetasks= {}
        #BUSINESS_RULES: Get Tasks and process them into the excel sheet.
        if sheetname !=  wsClosedOperations:
            casetasks = servicenowRestAPI.getCaseswithTasks(url_servicenow,item.get("sys_id"),user,pwd)


        if len(casetasks) != 0:
            # add to next row under the case above
            logging.info('Case Tasks are assoicated with = ' + v_hyperlink  )
            

            print len(casetasks)
            print len(Waiting_for_Account_Tasks)

            for s in casetasks:
                print s
                rowStart = rowStart + 1

                _cellA = ws['A' + str(rowStart)]
                _cellA.style = highlights
                v_hyperlink = url_servicenow + '/nav_to.do?uri=sn_customerservice_case.do?sys_id=' + item.get("sys_id") + '%26sysparm_view=case'
                _cellA.hyperlink = v_hyperlink
                _cellA.value = item.get("number") + '| \n' + s.get("number").get("display_value")
                _cellA.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

                _cellC = ws['B' + str(rowStart)]
                _cellC.style =  highlights
                _cellC.value = item.get("short_description") + ' | \n' + s.get("short_description").get("display_value")
                _cellC.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

                _cellD = ws['C' + str(rowStart)]
                _cellD.style =  highlights
                pValue = s.get("priority").get("display_value")
                pValue = pValue[3:]
                _cellD.value = pValue
                
                _cellD.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

                _cellE = ws['D' + str(rowStart)]
                _cellE.style =  highlights
                _cellE.value = s.get("state").get("display_value")
                _cellE.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

                sys_created_on_value = datetime.strptime(s.get("sys_created_on").get("display_value"), "%Y-%m-%d %H:%M:%S")
                _cellD = ws['E' + str(rowStart)]
                _cellD.style =  highlights
                _cellD.value = sys_created_on_value
                _cellD.number_format = 'MM/DD/YYYY'
                _cellD.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

                if isinstance(s.get("due_date").get("display_value"), datetime) == 'True':
                    due_date_value = datetime.strptime(s.get("due_date").get("display_value"), "%Y-%m-%d %H:%M:%S")
                else:
                    due_date_value = ''

                due_date_value = due_date_value
                _cellF = ws['F' + str(rowStart)]
                _cellF.style =  getdatetimeColorFormat(due_date_value,highlights)
                _cellF.value =  due_date_value
                _cellF.number_format = 'MM/DD/YYYY'
                _cellF.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

                _cellH = ws['G' + str(rowStart)]
                _cellH.style =  highlights
                _cellH.value = s.get("assigned_to").get("display_value")
                _cellH.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

                _cellI = ws['H' + str(rowStart)]
                _cellI.style = highlights
                _cellI.value = s.get("description").get("display_value")
                _cellI.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)

                _cellJ = ws['I' + str(rowStart)]
                _cellJ.style =  highlights
                _cellJ.value =  ' '
                _cellJ.alignment = Alignment(wrap_text= True,horizontal='left',vertical='top',shrinkToFit= True)
                
                if s.get("state").get("display_value") == 'Awaiting info': 
                    print s
                    global Waiting_for_Account_Tasks
                    Waiting_for_Account_Tasks.append(s)


                #Increment row to start a new Task not a new case
                #rowStart = rowStart + 1

            #Increment row to start a new start row for cases and not tasks
            #rowStart = rowStart + 1

        #capure the row end
        rowend = rowStart

    if sheetname == wsClosedOperations:
        refTables = 'A' + str(rowStartHolder+1)+ ':J' +str(rowend)
    else:
        refTables = 'A' + str(rowStartHolder+1)+ ':I' +str(rowend)

    tableName = title.replace(' ','')


    tab = Table(displayName= tableName, ref= refTables)

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)


    rowend = rowend +2  #this gives space between tables in worksheets that need more tabs

    # COLUMN WIDTH

    ws.column_dimensions["A"].width = 14  # CASE LINK
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20 # create date
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12 # due date
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["I"].width = 40
    if sheetname ==  wsClosedOperations:
        ws.column_dimensions["J"].width = 40

    if sheetname == wsClientAction:
        ws.sheet_properties.tabColor = wsClientActionTabColor

    return rowend
