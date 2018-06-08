import sys
import openpyxl
import datetime
import re
import os
import datetime
import logging
import random
import urllib
import createNewSheet
import servicenowRestAPI

from datetime import datetime,timedelta,date
from openpyxl import load_workbook
from datetime import datetime
from operator import itemgetter, attrgetter
from openpyxl.worksheet.table import Table
from openpyxl.styles import NamedStyle, Font
from openpyxl.drawing.image import Image


wsPartnerOps = 'Partner Operations'
wsPartnerClosedOps = 'Partner Closed Operations'
wsPartnerSummary = 'Partner Summary'
rowsStartsPerPage = {}




def  checkifPartner(wb, accountname,newFilepath, url_servicenow,account_sys_id,user,pwd):

    #GET Acounts if they have a parent
    allAccounts = servicenowRestAPI.getParentAccount(url_servicenow,account_sys_id,user,pwd)
    logging.info('checkpartner')


    rowsStartsPerPage.update({'wsPartnerOps': 4})
    rowsStartsPerPage.update({'wsPartnerClosedOps': 4})
    rowsStartsPerPage.update({'wsPartnerSummary': 4})

    if len(allAccounts) != 0:
        if wsPartnerOps not in wb.sheetnames:
            logging.info('createsheet: ' + wsPartnerOps)
            createNewSheet.createPlanningWSHeader(wb,wsPartnerOps,accountname, '' ,'',4,7)
        if wsPartnerClosedOps not in wb.sheetnames:
            createNewSheet.createPlanningWSHeader(wb,wsPartnerClosedOps,accountname, '' ,'',4,7)
            logging.info('createsheet: ' + wsPartnerClosedOps)
        if wsPartnerSummary not in wb.sheetnames:
            createNewSheet.createPlanningWSHeader(wb,wsPartnerSummary,accountname, '' ,'',4,7)
            logging.info('createsheet: ' + wsPartnerSummary)

        for acct in allAccounts:
            logging.info('proceessing account: ' + acct.get("sys_id"))
            acct_sys_id = acct.get("sys_id")

            caselist = servicenowRestAPI.getcasesTOupdateExcel(url_servicenow,user,pwd,account_sys_id)

            updateExcelwithPartnercases(wb,newFilepath,caselist,accountname,account_sys_id)

    return True


def updateExcelwithPartnercases(wb,newFilepath,caselist,accountname,account_sys_id,user,pwd):


    #TODO take list and make sublists to post into excel

    closed_cases = []
    Account_Planning = []
    Consumer_Inquiry = []
    Account_Strategy = []
    Account_Request = []
    Customer_Inquiry = []
    Account_Objective = []
    Waiting_for_Account = []
    left_overs = []  # Should be empty

    for s in caselist:


        if  s.get("active") == "false":
            closed_cases.append(s)
        elif s.get("category") == "Account Planning":
            Account_Planning.append(s)
        elif s.get("category") == "Account Requests":
            Account_Request.append(s)
        elif s.get("category") == "Consumer Inquiry":
            Consumer_Inquiry.append(s)
        elif s.get("category") == "Account Strategy":
            Account_Strategy.append(s)
        elif s.get("category") == "Customer Inquiry":
            Customer_Inquiry.append(s)
        elif s.get("category") == "Account Objective":
            Account_Objective.append(s)

        if s.get("state") == "Awaiting Info":
            Waiting_for_Account.append(s)


    #Account Planning
    endrow =  rowsStartsPerPage['wsPartnerOps'].value  # for starting after the title
    if len(Account_Planning) != 0:
        endrow  =  createNewSheet.section_table(accountname, Account_Planning,'Account Planning',endrow,wb,wsPartnerOps,user,pwd)

    if len(Consumer_Inquiry) != 0:
        endrow  =  createNewSheet.section_table(accountname,Consumer_Inquiry,'Consumer Inquiry',endrow,wb,wsPartnerOps,user,pwd)


    if len(Customer_Inquiry) != 0:
        endrow  =  createNewSheet.section_table(accountname,Customer_Inquiry,'Customer Inquiry',endrow,wb,wsPartnerOps,user,pwd)
    if len(left_overs) != 0:
        endrow  =  createNewSheet.section_table(accountname,left_overs,'Other Cases',endrow,wb,wsPartnerOps,user,pwd)

    #UPDATE ROW STARTS
    rowsStartsPerPage['wsPartnerOps'] = endrow

    #Start Closed Operatations sheet
    endrow = rowsStartsPerPage['wsPartnerClosedOps'].value
    if len(closed_cases) != 0:
        endrow  =  createNewSheet.section_table(accountname,closed_cases,'Closed Operations',endrow,wb,wsPartnerClosedOps,user,pwd)
        #UPDATE ROW STARTS
        rowsStartsPerPage['wsPartnerClosedOps'] = endrow

    #Start Account Summary Sheet
    endrow = rowsStartsPerPage['wsPartnerSummary'].value # for starting after the title
    if len(Waiting_for_Account) != 0:
        endrow  =  createNewSheet.section_table(accountname,Waiting_for_Account,'Account Summary',endrow,wb,wsPartnerSummary,user,pwd)
        #UPDATE ROW STARTS
        rowsStartsPerPage['wsPartnerSummary'] = endrow


    return wb







































































